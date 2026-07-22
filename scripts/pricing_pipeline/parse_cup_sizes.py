"""
Parses the cup-size (ounces) comparison side-table on the Frozen Yogurt
sheet (columns I:O, separate from the priced product rows in columns A:H) —
a per-brand, per-size-tier ounces comparison, not a price. Output feeds a
dedicated chart on the dashboard, not the price analytics pipeline.
"""

import openpyxl

SIZE_COLUMNS = {"S": "S", "M": "M", "L": "L", "TO GO": "TO_GO", "DINE IN": "DINE_IN"}


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _clean_ounces(value):
    value = _clean(value)
    if value is None or value == "-":
        return None
    if isinstance(value, str):
        # Normalize a double-dash range ("3--5") to a single-dash range ("3-5").
        return value.replace("--", "-")
    return value


def parse_cup_size_table(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=14, values_only=True))

    header_row_index = None
    for i, row in enumerate(rows):
        if _clean(row[0]) == "OUNCES":
            header_row_index = i
            break
    if header_row_index is None:
        raise ValueError("Could not locate the 'OUNCES' header in the cup-size table (columns I:N).")

    header = rows[header_row_index]
    column_keys = [None]  # column 0 is the brand name, not a size
    for cell in header[1:]:
        cleaned = _clean(cell)
        column_keys.append(SIZE_COLUMNS.get(cleaned))

    brand_aliases = {_clean(k): _clean(v) for k, v in config.get("cup_size_brand_aliases", {}).items()}

    result_rows = []
    for row in rows[header_row_index + 1:]:
        raw_brand = _clean(row[0])
        if raw_brand is None:
            break
        brand = brand_aliases.get(raw_brand, raw_brand)
        sizes = {}
        for col_index, key in enumerate(column_keys):
            if key is None:
                continue
            sizes[key] = _clean_ounces(row[col_index])
        result_rows.append({"brand": brand, **sizes})

    return {
        "meta": {"client": config["client"], "generated_from": xlsx_path},
        "sizes": ["S", "M", "L", "TO_GO", "DINE_IN"],
        "rows": result_rows,
    }


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_cup_size_table(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse the Frozen Yogurt cup-size comparison table into JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
