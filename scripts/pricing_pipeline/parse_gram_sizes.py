"""
Parses the gram-weight comparison side-table on the Frozen Yogurt sheet
(columns I:M, directly below the ounces table — headed "ML" in the client's
own file, though every value carries a "g" suffix, so it's grams, not
millilitres) — a per-brand, per-size-tier weight comparison, not a price.
Output feeds a dedicated chart on the dashboard, not the price analytics
pipeline.
"""

import openpyxl

SIZE_COLUMNS = {"S": "S", "M": "M", "L": "L", "Family": "FAMILY"}


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _clean_grams(value):
    value = _clean(value)
    if value is None or value == "-":
        return None
    if isinstance(value, str) and value.lower().endswith("g"):
        number = value[:-1].strip()
        try:
            return int(number)
        except ValueError:
            return value
    return value


def parse_gram_size_table(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=13, values_only=True))

    header_row_index = None
    for i, row in enumerate(rows):
        if _clean(row[0]) == "ML":
            header_row_index = i
            break
    if header_row_index is None:
        raise ValueError("Could not locate the 'ML' header in the gram-size table (columns I:M).")

    header = rows[header_row_index]
    column_keys = [None]  # column 0 is the brand name, not a size
    for cell in header[1:]:
        cleaned = _clean(cell)
        column_keys.append(SIZE_COLUMNS.get(cleaned))

    brand_aliases = {_clean(k): _clean(v) for k, v in config.get("cup_size_brand_aliases", {}).items()}

    result_rows = []
    for row in rows[header_row_index + 1:]:
        raw_brand = _clean(row[0])
        if raw_brand is None:
            break
        brand = brand_aliases.get(raw_brand, raw_brand)
        sizes = {}
        for col_index, key in enumerate(column_keys):
            if key is None:
                continue
            sizes[key] = _clean_grams(row[col_index])
        result_rows.append({"brand": brand, **sizes})

    return {
        "meta": {"client": config["client"], "generated_from": xlsx_path},
        "sizes": ["S", "M", "L", "FAMILY"],
        "rows": result_rows,
    }


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_gram_size_table(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse the Frozen Yogurt gram-size comparison table into JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
