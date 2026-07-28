"""
Parses the portion-size (grams) comparison side-table on the Salads sheet
(a 3-column block located by its "PORTION SIZE (g)" header, wherever the
main priced-product table happens to end — that boundary has already
shifted once as competitor columns were added) — a per-brand record of
what portion size the listed price is based on, and what alternate size
(if any) the brand also sells. Output feeds a dedicated disclosure table
on the dashboard, not the price analytics pipeline.
"""

import openpyxl

SIZE_COLUMNS = {"PRICED PORTION (G)": "PRICED_G", "ALSO AVAILABLE (G)": "ALSO_AVAILABLE_G"}


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _clean_grams(value):
    value = _clean(value)
    if value is None or value == "-":
        return None
    return value


def parse_portion_size_table(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    full_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    header_row_index = None
    header_col_index = None
    for i, row in enumerate(full_rows):
        for j, cell in enumerate(row):
            if _clean(cell) == "PORTION SIZE (g)":
                header_row_index, header_col_index = i, j
                break
        if header_col_index is not None:
            break
    if header_col_index is None:
        raise ValueError("Could not locate the 'PORTION SIZE (g)' header anywhere in the sheet.")

    rows = [row[header_col_index:header_col_index + 3] for row in full_rows]
    header = rows[header_row_index]
    column_keys = [None]  # column 0 is the brand name, not a size
    for cell in header[1:]:
        cleaned = _clean(cell)
        column_keys.append(SIZE_COLUMNS.get(cleaned))

    brand_aliases = {_clean(k): _clean(v) for k, v in config.get("portion_size_brand_aliases", {}).items()}
    dropped_brands = {_clean(b) for b in config.get("dropped_brands", [])}

    result_rows = []
    for row in rows[header_row_index + 1:]:
        raw_brand = _clean(row[0])
        if raw_brand is None:
            break
        brand = brand_aliases.get(raw_brand, raw_brand)
        if brand in dropped_brands:
            continue
        sizes = {}
        for col_index, key in enumerate(column_keys):
            if key is None:
                continue
            sizes[key] = _clean_grams(row[col_index])
        result_rows.append({"brand": brand, **sizes})

    return {
        "meta": {"client": config["client"], "generated_from": xlsx_path},
        "sizes": ["PRICED_G", "ALSO_AVAILABLE_G"],
        "rows": result_rows,
    }


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_portion_size_table(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse the Salads portion-size comparison table into JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
