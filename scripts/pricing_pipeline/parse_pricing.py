def is_category_header_row(row: tuple, brand_span_end: int = 8) -> bool:
    """A category header row has a product/category name in column 0 and
    nothing in any brand or portion-note column. brand_span_end must cover
    every brand column (up to but excluding the Average column) — a fixed
    window here previously missed products priced only by a brand sitting
    past column 7 (e.g. a 5th competitor), silently misreading them as
    category headers and dropping them entirely."""
    product = row[0]
    rest = row[1:brand_span_end]
    return product is not None and all(v is None for v in rest)


import openpyxl


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def parse_workbook(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    header = rows[0]

    # Locate the average column to determine the actual brand column span.
    # Most sheets label it "Average"; the Non-Dairy comparison file labels
    # it "Avg" instead — accept both rather than adding a config key for
    # what is a structural header-label variant, not client-specific data.
    average_col_index = None
    for i, col_header in enumerate(header):
        if _clean(col_header) in ("Average", "Avg"):
            average_col_index = i
            break

    if average_col_index is None:
        raise ValueError(
            "Could not locate the 'Average' column in the header row — "
            "sheet layout does not match the expected template."
        )

    # Extract the FULL actual set of brand columns from the file, then
    # translate any header text an alias covers (e.g. "Esp. Lab" -> "Espresso
    # Lab") to the canonical brand name before validating against config.
    # Brand columns are not always contiguous — some sheets interleave an
    # unlabeled per-product portion-note column ahead of a brand's price
    # column, so only actual header cells (skipping the blank ones) count.
    # dropped_brands are present in the sheet (kept there as reference
    # information the client wants on file) but deliberately excluded from
    # the report — e.g. a competitor superseded by a newly-added one.
    brand_aliases = {_clean(k): _clean(v) for k, v in config.get("brand_aliases", {}).items()}
    dropped_brands = {_clean(b) for b in config.get("dropped_brands", [])}
    all_col_indices = [i for i in range(1, average_col_index) if _clean(header[i]) is not None]
    all_translated = [brand_aliases.get(_clean(header[i]), _clean(header[i])) for i in all_col_indices]
    brand_col_indices = [i for i, b in zip(all_col_indices, all_translated) if b not in dropped_brands]
    brand_columns = [b for b in all_translated if b not in dropped_brands]

    # Validate that header brands match config brands (bidirectional check)
    expected_brands = {_clean(config["own_brand"]), *[_clean(c) for c in config["competitors"]]}
    actual_brands = set(brand_columns)

    if actual_brands != expected_brands:
        missing = expected_brands - actual_brands
        extra = actual_brands - expected_brands
        raise ValueError(
            f"Header brand columns {brand_columns} do not match config's expected brands "
            f"(own_brand + competitors). Missing from header: {sorted(missing) or 'none'}. "
            f"Present in header but not in config: {sorted(extra) or 'none'}."
        )

    dropped_categories = {_clean(c) for c in config.get("dropped_categories", [])}
    category_aliases = {_clean(k): _clean(v) for k, v in config.get("category_aliases", {}).items()}

    fx_rate = config["fx_usd_rate"]
    records = []
    unparseable_prices = []
    current_category = None

    for row in rows[1:]:
        product = _clean(row[0])
        if product is None:
            continue

        if is_category_header_row(row, average_col_index):
            current_category = category_aliases.get(product, product)
            continue

        if current_category in dropped_categories:
            continue

        for idx, brand in zip(brand_col_indices, brand_columns):
            price = row[idx]
            if not isinstance(price, (int, float)):
                # "-" and blank cells are the sheet's normal way of saying
                # "no price here" — anything else non-numeric (e.g. a comma-
                # decimal typo like "8000,00") is a real data-quality problem
                # that was previously dropped with no trace. Disclose it
                # instead of guessing at the intended value.
                cleaned_price = _clean(price)
                if cleaned_price not in (None, "-"):
                    unparseable_prices.append({
                        "category": current_category,
                        "product": product,
                        "brand": brand,
                        "raw_value": cleaned_price,
                    })
                continue

            # An unlabeled column immediately to the left of a brand's price
            # column (present on some sheets, e.g. Salads) carries that
            # brand's specific portion size for this product — text like
            # "380 g." — rather than a price.
            portion_note = None
            if _clean(header[idx - 1]) is None:
                note_value = _clean(row[idx - 1])
                if note_value not in (None, "-"):
                    portion_note = note_value

            records.append({
                "category": current_category,
                "product": product,
                "brand": brand,
                "price_lbp": price,
                "price_usd": round(price / fx_rate, 2),
                "portion_note": portion_note,
            })

    meta = {
        "client": config["client"],
        "report_date": config["report_date"],
        "currency": config["currency"],
        "fx_usd_rate": fx_rate,
        "fx_rate_date": config["fx_rate_date"],
        "fx_source": config["fx_source"],
        "own_brand": _clean(config["own_brand"]),
        "competitors": [_clean(c) for c in config["competitors"]],
        "generated_from": xlsx_path,
    }

    return {"meta": meta, "records": records, "unparseable_prices": unparseable_prices}


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_workbook(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse a raw pricing-comparison Excel file into normalized JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the normalized JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
