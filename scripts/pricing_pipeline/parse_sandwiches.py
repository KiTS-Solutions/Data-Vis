"""
Dedicated parser for the Sandwich comparison file, whose layout doesn't
match the standard template parse_pricing.parse_workbook expects: a
two-row title/subtitle preamble before the header, product identity split
across two columns (Item + Bread Type), derived summary columns (Min/Max/
Average/Lowest Bakery) to ignore, a trailing summary row, and a second,
unrelated "Bakery Insights" sheet. Rather than bend the generic parser to
one file's shape, this reads it directly into the same normalized record
format every other source produces, so analyze_pricing.py needs no changes.
"""

import openpyxl

SHEET_NAME = "Price Comparison"
PRICE_COLUMN_SUFFIX = " (LBP)"


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def parse_sandwich_workbook(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Locate the header row by content, not a fixed offset — the preamble
    # above it (title, subtitle, blank separator rows) isn't a fixed count.
    header_row_index = None
    for i, row in enumerate(rows):
        if _clean(row[0]) == "Item":
            header_row_index = i
            break
    if header_row_index is None:
        raise ValueError("Could not locate the 'Item' header row in the Sandwich sheet.")

    header = rows[header_row_index]
    brand_columns = []  # list of (column_index, brand_name)
    for i, col_header in enumerate(header):
        cleaned = _clean(col_header)
        if isinstance(cleaned, str) and cleaned.endswith(PRICE_COLUMN_SUFFIX):
            brand_columns.append((i, cleaned[: -len(PRICE_COLUMN_SUFFIX)]))

    expected_brands = {_clean(config["own_brand"]), *[_clean(c) for c in config["competitors"]]}
    actual_brands = {b for _, b in brand_columns}
    if actual_brands != expected_brands:
        missing = expected_brands - actual_brands
        extra = actual_brands - expected_brands
        raise ValueError(
            f"Sandwich sheet brand columns {sorted(actual_brands)} do not match config's expected "
            f"brands. Missing from sheet: {sorted(missing) or 'none'}. "
            f"Present in sheet but not in config: {sorted(extra) or 'none'}."
        )

    category = config["category"]
    fx_rate = config["fx_usd_rate"]
    records = []

    for row in rows[header_row_index + 1:]:
        item = _clean(row[0])
        if item is None or item == "Average Price":
            continue
        bread_type = _clean(row[1])
        product = f"{item} ({bread_type})" if bread_type else item

        for col_index, brand in brand_columns:
            price = row[col_index]
            if not isinstance(price, (int, float)):
                continue
            records.append({
                "category": category,
                "product": product,
                "brand": brand,
                "price_lbp": price,
                "price_usd": round(price / fx_rate, 2),
            })

    meta = {
        "client": config["client"],
        "report_date": config["report_date"],
        "currency": config["currency"],
        "fx_usd_rate": fx_rate,
        "fx_rate_date": config["fx_rate_date"],
        "fx_source": config["fx_source"],
        "own_brand": _clean(config["own_brand"]),
        "competitors": [_clean(c) for c in config["competitors"]],
        "generated_from": xlsx_path,
    }

    return {"meta": meta, "records": records}


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_sandwich_workbook(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse the Sandwich comparison Excel file into normalized JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the normalized JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
