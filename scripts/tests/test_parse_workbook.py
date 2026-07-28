import openpyxl
from pricing_pipeline.parse_pricing import parse_workbook


def _build_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts",
               "Joe & the Juice", "Starbucks ", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None, None, None])
    ws.append(["Double Espresso Macchiato", 300000, "-", "-", "-", "-", 300000, 0])
    ws.append(["Americano MEDIUM", 350000, 400000, "-", 358000, 350000, 364500, -14500])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


def _config():
    return {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts", "Joe & the Juice", "Starbucks"],
    }


def test_parse_workbook_skips_dashes_and_assigns_category(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    macchiato = [r for r in result["records"] if r["product"] == "Double Espresso Macchiato"]
    assert len(macchiato) == 1
    assert macchiato[0]["brand"] == "Stories"
    assert macchiato[0]["price_lbp"] == 300000
    assert macchiato[0]["category"] == "Black Coffee"
    assert macchiato[0]["price_usd"] == 3.35


def test_parse_workbook_multiple_brands_per_product(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    americano = [r for r in result["records"] if r["product"] == "Americano MEDIUM"]
    assert len(americano) == 4
    assert {r["brand"] for r in americano} == {"Stories", "Espresso Lab", "Joe & the Juice", "Starbucks"}


def test_parse_workbook_meta_fields(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    assert result["meta"]["client"] == "Stories"
    assert result["meta"]["own_brand"] == "Stories"
    assert result["meta"]["generated_from"] == xlsx_path


def test_parse_workbook_different_competitor_count(tmp_path):
    """Test that brand column count is derived from config, not hardcoded to 5."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 2 competitors = 3 brand columns total (1 own + 2 competitors)
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None])
    ws.append(["Americano MEDIUM", 350000, 400000, 380000, 376666, -26666])
    path = tmp_path / "sample_2competitors.xlsx"
    wb.save(path)

    config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts"],  # Only 2 competitors
    }

    result = parse_workbook(str(path), config)
    americano = [r for r in result["records"] if r["product"] == "Americano MEDIUM"]
    assert len(americano) == 3
    assert {r["brand"] for r in americano} == {"Stories", "Espresso Lab", "Dunkin Donuts"}


def test_parse_workbook_config_mismatch_raises_error(tmp_path):
    """Test that mismatched config and header brands raise ValueError."""
    xlsx_path = _build_workbook(tmp_path)

    # Config with a competitor not in the header
    bad_config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts", "Unknown Brand", "Starbucks"],
    }

    try:
        parse_workbook(xlsx_path, bad_config)
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "Unknown Brand" in str(e)
        assert "do not match config's expected brands" in str(e)


def test_parse_workbook_header_has_extra_brands_not_in_config(tmp_path):
    """
    Test that raises ValueError when header has MORE brands than config expects.
    This reproduces the scenario where someone adds a 5th competitor to the
    spreadsheet (e.g., "E") but forgets to update the config.
    Before the fix, this would silently drop the extra column with no error.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Header has 5 brands (Stories, A, B, C, D, E) + Average + Difference
    ws.append(["Products Competitors", "Stories", "A", "B", "C", "D", "E", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None, None, None, None])
    ws.append(["Americano MEDIUM", 350000, 400000, 380000, 360000, 370000, 390000, 374333, 24333])
    path = tmp_path / "sample_extra_brands.xlsx"
    wb.save(path)

    # Config expects only 4 competitors (A, B, C, D) — missing E
    config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["A", "B", "C", "D"],  # Only 4, missing E
    }

    try:
        parse_workbook(str(path), config)
        assert False, "Should have raised ValueError when header has extra brands not in config"
    except ValueError as e:
        error_msg = str(e)
        assert "do not match config's expected brands" in error_msg
        assert "E" in error_msg
        assert "Present in header but not in config" in error_msg


def test_parse_workbook_skips_dropped_categories(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts",
               "Joe & the Juice", "Starbucks ", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None, None, None])
    ws.append(["Double Espresso Macchiato", 300000, "-", "-", "-", "-", 300000, 0])
    ws.append(["Salads", None, None, None, None, None, None, None])
    ws.append(["Quinoa Salad", 770000, "-", "-", "-", "-", 770000, 0])
    path = tmp_path / "sample_dropped.xlsx"
    wb.save(path)

    config = _config()
    config["dropped_categories"] = ["Salads"]

    result = parse_workbook(str(path), config)

    categories = {r["category"] for r in result["records"]}
    assert categories == {"Black Coffee"}
    assert all(r["product"] != "Quinoa Salad" for r in result["records"])


def test_parse_workbook_without_dropped_categories_key_is_unaffected(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())
    assert len(result["records"]) == 5


def test_parse_workbook_accepts_avg_as_average_header(tmp_path):
    """Some sheets label the average column 'Avg' instead of 'Average'
    (e.g. the real Non-Dairy Product Pricing Comparison file) — the
    average-column locator must accept both."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts",
               "Joe & the Juice", "Starbucks ", "Avg", "Dif."])
    ws.append(["Black Coffee", None, None, None, None, None, None, None])
    ws.append(["Double Espresso Macchiato", 300000, "-", "-", "-", "-", 300000, 0])
    path = tmp_path / "sample_avg_header.xlsx"
    wb.save(path)

    result = parse_workbook(str(path), _config())

    assert len(result["records"]) == 1
    assert result["records"][0]["price_lbp"] == 300000


def test_parse_workbook_applies_brand_aliases(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Esp. Lab", "D. Donuts",
               "J & J", "Starbucks ", "Avg", "Dif."])
    ws.append(["Mixed Hot Beverages", None, None, None, None, None, None, None])
    ws.append(["Cappuccino MEDIUM", 600000, 650000, 576000, 554900, 550000, 586180, 13820])
    path = tmp_path / "sample_aliased.xlsx"
    wb.save(path)

    config = _config()
    config["brand_aliases"] = {
        "Esp. Lab": "Espresso Lab",
        "D. Donuts": "Dunkin Donuts",
        "J & J": "Joe & the Juice",
    }

    result = parse_workbook(str(path), config)

    brands = {r["brand"] for r in result["records"]}
    assert brands == {"Stories", "Espresso Lab", "Dunkin Donuts", "Joe & the Juice", "Starbucks"}
    cappuccino_stories = [r for r in result["records"] if r["product"] == "Cappuccino MEDIUM" and r["brand"] == "Stories"]
    assert cappuccino_stories[0]["price_lbp"] == 600000


def test_parse_workbook_skips_unlabeled_portion_note_columns(tmp_path):
    """Reproduces the real Salads sheet: an unlabeled per-product portion-note
    column (e.g. "380 g.") sits ahead of most brands' price columns, but not
    every brand has one (the last competitor here has none) — brand columns
    must be located by their actual header index, not assumed contiguous."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", None, "Stories", None, "Wooden Bakery", None, "Zaatar w Zeit",
               "The Koozspace", "Average", "Difference"])
    ws.append(["SALADS", None, None, None, None, None, None, None, None, None])
    ws.append(["ASIAN SALAD", "380 g.", 750000, "-", "-", None, "-", "-", 750000, 0])
    ws.append(["TUNA PASTA SALAD", "440 g.", 850000, "550 g.", 716800, "200 g.", 806400, 1075200, 862100, -12100])
    path = tmp_path / "sample_gapped_columns.xlsx"
    wb.save(path)

    config = _config()
    config["competitors"] = ["Wooden Bakery", "Zaatar w Zeit", "The Koozspace"]

    result = parse_workbook(str(path), config)

    tuna = {r["brand"]: r["price_lbp"] for r in result["records"] if r["product"] == "TUNA PASTA SALAD"}
    assert tuna == {"Stories": 850000, "Wooden Bakery": 716800, "Zaatar w Zeit": 806400, "The Koozspace": 1075200}


def test_parse_workbook_captures_per_product_portion_notes(tmp_path):
    """The unlabeled column ahead of a brand's price column carries that
    brand's specific portion size for the product (e.g. "380 g."), not a
    price — it should be captured as portion_note, not read as a value."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", None, "Stories", None, "Wooden Bakery", "The Koozspace",
               "Average", "Difference"])
    ws.append(["SALADS", None, None, None, None, None, None, None])
    ws.append(["ASIAN SALAD", "380 g.", 750000, "-", "-", "-", 750000, 0])
    ws.append(["TUNA PASTA SALAD", "440 g.", 850000, "550 g.", 716800, 1075200, 862100, -12100])
    path = tmp_path / "sample_portion_notes.xlsx"
    wb.save(path)

    config = _config()
    config["competitors"] = ["Wooden Bakery", "The Koozspace"]

    result = parse_workbook(str(path), config)

    tuna_stories = next(r for r in result["records"] if r["product"] == "TUNA PASTA SALAD" and r["brand"] == "Stories")
    assert tuna_stories["portion_note"] == "440 g."

    tuna_wb = next(r for r in result["records"] if r["product"] == "TUNA PASTA SALAD" and r["brand"] == "Wooden Bakery")
    assert tuna_wb["portion_note"] == "550 g."

    tuna_koozspace = next(r for r in result["records"] if r["product"] == "TUNA PASTA SALAD" and r["brand"] == "The Koozspace")
    assert tuna_koozspace["portion_note"] is None

    # ASIAN SALAD has "-" (no price) for Wooden Bakery, so no record at all is created for it.
    assert not any(r["product"] == "ASIAN SALAD" and r["brand"] == "Wooden Bakery" for r in result["records"])


def test_parse_workbook_dropped_brands_excluded_from_records_and_validation(tmp_path):
    """A brand column can be present in the sheet (kept there as reference
    the client wants on file) without being in config.competitors, as long
    as it's explicitly named in dropped_brands — otherwise this would raise
    the "extra brand in header" validation error."""
    xlsx_path = _build_workbook(tmp_path)
    config = _config()
    config["competitors"] = ["Espresso Lab", "Dunkin Donuts", "Joe & the Juice"]  # Starbucks dropped
    config["dropped_brands"] = ["Starbucks"]

    result = parse_workbook(xlsx_path, config)

    macchiato = [r for r in result["records"] if r["product"] == "Double Espresso Macchiato"]
    assert {r["brand"] for r in macchiato} == {"Stories"}
    americano = [r for r in result["records"] if r["product"] == "Americano MEDIUM"]
    assert {r["brand"] for r in americano} == {"Stories", "Espresso Lab", "Joe & the Juice"}


def test_parse_workbook_applies_category_aliases(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts",
               "Joe & the Juice", "Starbucks ", "Average", "Difference"])
    ws.append(["SALADS", None, None, None, None, None, None, None])
    ws.append(["Caesar Salad", 750000, "-", "-", "-", "-", 750000, 0])
    ws.append(["Gap analysis items not on Stories menu", None, None, None, None, None, None, None])
    ws.append(["Crab Salad", "-", "-", "-", "-", 806400, 806400, "-"])
    path = tmp_path / "sample_category_alias.xlsx"
    wb.save(path)

    config = _config()
    config["category_aliases"] = {"Gap analysis items not on Stories menu": "SALADS"}

    result = parse_workbook(str(path), config)

    categories = {r["category"] for r in result["records"]}
    assert categories == {"SALADS"}
    assert any(r["product"] == "Crab Salad" for r in result["records"])
